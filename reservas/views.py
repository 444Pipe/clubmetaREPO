from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Sum, Count
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import logout
from django.contrib import messages
from .models import Salon, ConfiguracionSalon, Reserva, BloqueoEspacio
import csv
from datetime import datetime
import datetime as dt

# Mapeo de imágenes por salón (hardcoded por ahora)
SALON_IMAGES = {
    'Mi Llanura': [
        "salon mi llanura/WhatsApp Image 2025-12-01 at 11.51.19 AM (1).jpeg",
        "salon mi llanura/WhatsApp Image 2025-12-01 at 11.51.19 AM.jpeg",
        "salon mi llanura/WhatsApp Image 2025-12-01 at 11.51.20 AM.jpeg",
        "salon mi llanura/milanuraa.avif",
        "salon mi llanura/millanura.avif",
        "salon mi llanura/millanuraaa.avif",
        "salon mi llanura/mika.jpg"
    ],
    'Salón Bar': [
        "salon bar/bar.jpg",
        "salon bar/barsito.jpg"
    ],
    'Salón Empresarial': [
        "salon empresarial/salonempresarial.jpg",
        "salon empresarial/salonempresarialxd.jpg",
        "salon empresarial/WhatsApp Image 2025-12-01 at 11.50.32 AM (1).jpeg",
        "salon empresarial/WhatsApp Image 2025-12-01 at 11.50.32 AM.jpeg"
    ],
    'Terraza': [
        "salon terraza/terraza.jpg",
        "salon terraza/WhatsApp Image 2025-12-01 at 11.51.01 AM (1).jpeg",
        "salon terraza/WhatsApp Image 2025-12-01 at 11.51.01 AM.jpeg",
        "salon terraza/1.jpg"
    ],
    'Salón Kiosco': [
        "salon kiosco/kiosco.jpg",
        "salon kiosco/kioscoxdd.jpg",
        "salon kiosco/kioskoxd.jpg",
        "salon kiosco/kioskito.jpg",
        "salon kiosco/kioskito2.jpg",
        "salon kiosco/kiskito1.jpg",
    ],
    'Salón Presidente': ["salon presidente/presidente.jpg", "salon presidente/presidente1.jpg"]
}

def get_salon_images(salon_or_name):
    """Obtiene las imágenes del salón.

    Acepta tanto el nombre del salón (str) como una instancia de `Salon`.
    - Si se pasa una instancia y tiene el campo `imagen` no vacío, se devuelve esa ruta.
    - Luego intenta hacer match con las claves de `SALON_IMAGES` por nombre.
    - Si no encuentra nada devuelve lista vacía.
    """
    # Si nos pasan una instancia de Salon
    try:
        # detectar si es objeto con atributo 'imagen' y 'nombre'
        imagen_attr = getattr(salon_or_name, 'imagen', None)
        nombre_attr = getattr(salon_or_name, 'nombre', None)
    except Exception:
        imagen_attr = None
        nombre_attr = None

    # Si la instancia tiene una ruta en el campo imagen, úsala (como único elemento)
    if imagen_attr:
        return [imagen_attr]

    # Determinar el nombre a usar para buscar en SALON_IMAGES
    nombre_buscar = ''
    if nombre_attr:
        nombre_buscar = nombre_attr
    elif isinstance(salon_or_name, str):
        nombre_buscar = salon_or_name

    for key in SALON_IMAGES.keys():
        if key in (nombre_buscar or ''):
            return SALON_IMAGES[key]

    return []

# Validación de código de socio contra base de datos
def is_valid_socio(code):
    """Valida si el código de socio existe y está activo"""
    if not code:
        return False
    from .models import CodigoSocio
    return CodigoSocio.objects.filter(codigo=code.strip(), activo=True).exists()

def validate_socio_code(request):
    """Vista AJAX para validar código de socio en tiempo real"""
    if request.method == 'GET':
        code = request.GET.get('code', '').strip()
        if not code:
            return JsonResponse({'valid': False, 'message': 'Código vacío'})
        
        is_valid = is_valid_socio(code)
        if is_valid:
            # además de validar, devolver datos del socio para autocompletar el formulario
            from .models import CodigoSocio
            try:
                socio = CodigoSocio.objects.filter(codigo=code.strip(), activo=True).values(
                    'nombre_socio', 'email', 'celular', 'empresa'
                ).first()
            except Exception:
                socio = None

            return JsonResponse({'valid': True, 'message': 'Código válido ✓', 'socio': socio})
        else:
            return JsonResponse({'valid': False, 'message': 'Código inválido o inactivo'})
    
    return JsonResponse({'valid': False, 'message': 'Método no permitido'})

def index(request):
    """Página principal - bienvenida"""
    return render(request, 'index.html')


def preguntas_frecuentes(request):
    return render(request, 'preguntas_frecuentes.html')


def politicas(request):
    return render(request, 'politicas.html')

def espacios(request):
    """Página de espacios disponibles - usando datos de la base de datos"""
    from .models import ConfiguracionSalon
    
    salones = Salon.objects.filter(disponible=True).order_by('nombre')
    
    # Convertir salones a formato compatible con template
    # Ahora mostramos UN salón con sus configuraciones como opciones
    rooms = []
    for salon in salones:
        # Obtener todas las configuraciones del salón
        configuraciones = salon.configuraciones.all()
        
        if configuraciones.exists():
            # Usar la primera configuración para mostrar precio/capacidad de referencia
            primera_config = configuraciones.first()
            
            rooms.append({
                'id': salon.id,
                'name': salon.nombre,
                'desc': salon.descripcion,
                'price_particular_4h': int(primera_config.precio_particular_4h),
                'price_particular_8h': int(primera_config.precio_particular_8h or primera_config.precio_particular_4h),
                'price_socio_4h': int(primera_config.precio_socio_4h),
                'price_socio_8h': int(primera_config.precio_socio_8h or primera_config.precio_socio_4h),
                'capacity': primera_config.capacidad,
                'type': 'social',
                'images': get_salon_images(salon),
                'configuraciones': list(configuraciones)  # Lista de todas las configuraciones
            })
    
    return render(request, 'espacios.html', {'rooms': rooms})

def register(request):
    """Página de registro/reserva"""
    from .models import ConfiguracionSalon, ServicioAdicional
    
    configuraciones = ConfiguracionSalon.objects.filter(salon__disponible=True).select_related('salon').order_by('salon__nombre', 'tipo_configuracion')
    
    # Obtener servicios adicionales activos (excluir montajes ya que se manejan por separado)
    servicios_adicionales = ServicioAdicional.objects.filter(activo=True).exclude(nombre__icontains='montaje').order_by('nombre')
    
    # Convertir configuraciones a formato compatible con template
    rooms = []
    for config in configuraciones:
        # Mostrar solo precio particular (ocultar precios de socio por solicitud del cliente)
        precio_texto = f"${config.precio_particular_4h:,.0f}"
        
        rooms.append({
            'id': config.id,
            'salon_id': config.salon.id,
            'name': f"{config.salon.nombre} - {config.get_tipo_configuracion_display()}",
            'desc': f"{config.capacidad} personas",
            'price': precio_texto,  # Texto completo del precio
            'price_particular_4h': int(config.precio_particular_4h),
            'price_particular_8h': int(config.precio_particular_8h or config.precio_particular_4h),
            'price_socio_4h': int(config.precio_socio_4h),
            'price_socio_8h': int(config.precio_socio_8h or config.precio_socio_4h),
            'capacity': config.capacidad,
            'type': 'social',
            'images': get_salon_images(config.salon)
        })
    
    if request.method == 'GET':
        return render(request, 'register.html', {
            'rooms': rooms,
            'servicios_adicionales': servicios_adicionales
        })
    
    # Procesar POST
    nombre_cliente = request.POST.get('nombre_cliente', '').strip()
    email_cliente = request.POST.get('email_cliente', '').strip()
    telefono_cliente = request.POST.get('telefono_cliente', '').strip()
    num_personas = request.POST.get('num_personas', '').strip()
    fecha_evento = request.POST.get('fecha_evento', '').strip()
    duracion_horas = request.POST.get('duracion_horas', '').strip()
    hora_inicio = request.POST.get('hora_inicio', '').strip()
    tiempo_decoracion = request.POST.get('tiempo_decoracion', '0').strip()
    espacio_id = request.POST.get('espacio_id', '0')
    socio_code = request.POST.get('socio_code', '').strip()
    note = request.POST.get('note', '').strip()
    
    # Validaciones
    errors = []
    if not nombre_cliente or len(nombre_cliente) < 2:
        errors.append("Nombre del cliente inválido.")
    if not email_cliente or '@' not in email_cliente:
        errors.append("Email inválido.")
    if not telefono_cliente or len(telefono_cliente) < 7:
        errors.append("Teléfono inválido.")
    
    try:
        personas_i = int(num_personas)
        if personas_i <= 0:
            errors.append("Número de personas inválido.")
    except:
        errors.append("Número de personas inválido.")
        personas_i = 0
    
    try:
        duracion_i = int(duracion_horas)
        if duracion_i <= 0:
            errors.append("Duración del evento inválida.")
    except:
        errors.append("Duración del evento inválida.")
        duracion_i = 0
    
    # Validar fecha del evento
    try:
        from datetime import date
        fecha_evento_obj = datetime.strptime(fecha_evento, '%Y-%m-%d').date()
        if fecha_evento_obj < date.today():
            errors.append("La fecha del evento no puede ser en el pasado.")
    except ValueError:
        errors.append("Formato de fecha inválido.")
        fecha_evento_obj = None
    
    try:
        from .models import ConfiguracionSalon, BloqueoEspacio
        espacio_id_int = int(espacio_id)
        configuracion = ConfiguracionSalon.objects.filter(id=espacio_id_int, salon__disponible=True).select_related('salon').first()
        if not configuracion:
            errors.append("Espacio seleccionado inválido.")
        else:
            # Validar que el salón no esté bloqueado en esa fecha
            if fecha_evento_obj:
                bloqueos = BloqueoEspacio.objects.filter(
                    salon=configuracion.salon,
                    activo=True,
                    fecha_inicio__lte=fecha_evento_obj,
                    fecha_fin__gte=fecha_evento_obj
                )
                if bloqueos.exists():
                    bloqueo = bloqueos.first()
                    fecha_inicio_str = bloqueo.fecha_inicio.strftime('%d/%m/%Y')
                    fecha_fin_str = bloqueo.fecha_fin.strftime('%d/%m/%Y')
                    errors.append(f"El salón {configuracion.salon.nombre} no está disponible para la fecha {fecha_evento}. Motivo: {bloqueo.get_motivo_display()}. Bloqueado desde {fecha_inicio_str} hasta {fecha_fin_str}.")
            
            # Validar capacidad del salón
            if personas_i > configuracion.capacidad:
                errors.append(f"El salón soporta máximo {configuracion.capacidad} personas. Solicitaste {personas_i}.")
            
            # Validar código de socio si se marcó como socio
            es_socio = request.POST.get('es_socio', 'no')
            if es_socio == 'si' and not socio_code:
                errors.append("Debes ingresar el código de socio.")
            elif es_socio == 'si' and not is_valid_socio(socio_code):
                errors.append("El código de socio ingresado no es válido (mínimo 3 caracteres).")
    except:
        errors.append("Espacio seleccionado inválido.")
        configuracion = None
    
    if errors:
        return render(request, 'register.html', {'rooms': rooms, 'errors': errors, 'form': request.POST}, status=400)
    
    # Determine price depending on socio code and duration (4H/8H)
    es_socio = request.POST.get('es_socio', 'no')
    if configuracion:
        applied_socio = False
        if es_socio == 'si' and is_valid_socio(socio_code):
            applied_socio = True

        # Precio según duración seleccionada
        if duracion_i == 8:
            if applied_socio:
                price = int(configuracion.precio_socio_8h or configuracion.precio_socio_4h)
            else:
                price = int(configuracion.precio_particular_8h or configuracion.precio_particular_4h)
        else:
            if applied_socio:
                price = int(configuracion.precio_socio_4h)
            else:
                price = int(configuracion.precio_particular_4h)

        espacio_nombre = f"{configuracion.salon.nombre} - {configuracion.get_tipo_configuracion_display()}"
    else:
        price = 0
        applied_socio = False
        espacio_nombre = ''

    total_price = price
    
    # Procesar tiempo de decoración
    try:
        tiempo_decoracion_i = int(tiempo_decoracion)
    except:
        tiempo_decoracion_i = 0
    
    # Guardar la reserva en la base de datos
    try:
        from datetime import time as datetime_time
        # Convertir hora_inicio string a objeto time
        hora_inicio_obj = None
        if hora_inicio:
            try:
                hora_parts = hora_inicio.split(':')
                hora_inicio_obj = datetime_time(int(hora_parts[0]), int(hora_parts[1]))
            except:
                hora_inicio_obj = None

        # Validar rango permitido: desde 08:30 hasta 02:00 (noche siguiente)
        # Permitimos horas entre 08:30 (08:30) y 02:00 (02:00 del día siguiente).
        if hora_inicio_obj:
            try:
                from datetime import time as datetime_time_check
                min_minutes = 8 * 60 + 30
                # Representaremos las horas posteriores a medianoche como hora+24h para comparar
                h = hora_inicio_obj.hour
                m = hora_inicio_obj.minute
                minutes = (h if h >= 8 else h + 24) * 60 + m
                max_minutes = (2 + 24) * 60  # 26:00 -> 02:00 siguiente día
                if not (minutes >= min_minutes and minutes <= max_minutes):
                    errors.append('Hora inválida. Horarios disponibles: 08:30 a.m. a 02:00 a.m. (siguiente día).')
            except Exception:
                errors.append('Error al validar la hora de inicio.')
        
        # Obtener nombre de entidad si viene de empresa
        nombre_entidad = request.POST.get('nombre_entidad', '').strip()
        
        reserva = Reserva.objects.create(
            configuracion_salon=configuracion,
            nombre_cliente=nombre_cliente,
            email_cliente=email_cliente,
            telefono_cliente=telefono_cliente,
            tipo_cliente='SOCIO' if applied_socio else 'PARTICULAR',
            nombre_entidad=nombre_entidad if nombre_entidad else None,
            fecha_evento=fecha_evento,
            hora_inicio=hora_inicio_obj,
            duracion='4H' if duracion_i == 4 else '8H',
            tiempo_decoracion=tiempo_decoracion_i,
            numero_personas=personas_i,
            precio_total=0,
            estado='PENDIENTE',
            observaciones=note if note else ''
        )
        # Después de crear la reserva, refrescar para obtener el precio calculado por el modelo
        reserva.refresh_from_db()

        # Guardar servicios adicionales
        from .models import ServicioAdicional, ReservaServicioAdicional
        servicios_total = 0
        for key in request.POST:
            if key.startswith('servicio_') and request.POST[key]:
                try:
                    servicio_id = int(key.replace('servicio_', ''))
                    cantidad = int(request.POST[key])
                    if cantidad > 0:
                        servicio = ServicioAdicional.objects.get(id=servicio_id, activo=True)
                        reserva_servicio = ReservaServicioAdicional.objects.create(
                            reserva=reserva,
                            servicio=servicio,
                            cantidad=cantidad,
                            precio_unitario=servicio.precio_unitario
                        )
                        servicios_total += float(reserva_servicio.subtotal)
                except (ValueError, ServicioAdicional.DoesNotExist):
                    pass
        
        # Actualizar precio total de la reserva si hay servicios
        if servicios_total > 0:
            reserva.precio_total = float(reserva.precio_total) + servicios_total
            reserva.save()

        # Obtener el precio final calculado en la reserva
        reserva.refresh_from_db()
        total_price = int(reserva.precio_total or 0)
        
        # Mensaje de éxito
        messages.success(request, f'¡Reserva #{reserva.id} creada exitosamente! Total: ${total_price} USD')
        
        # Preparar payload para mostrar al usuario
        abono_50 = int(total_price * 0.5)
        payload = {
            'reserva_id': reserva.id,
            'nombre_cliente': nombre_cliente,
            'email_cliente': email_cliente,
            'telefono_cliente': telefono_cliente,
            'num_personas': personas_i,
            'fecha_evento': fecha_evento,
            'hora_inicio': hora_inicio if hora_inicio else 'No especificada',
            'duracion_horas': duracion_i,
            'espacio_id': espacio_id_int,
            'espacio_nombre': espacio_nombre,
            'precio_por_evento': price,
            'total_price': total_price,
            'abono_50': f"{abono_50:,}",
            'socio_code': socio_code,
            'applied_socio': applied_socio,
            'note': note
        }
        
        return render(request, 'register.html', {'success': True, 'registration': payload, 'rooms': rooms})
    
    except Exception as e:
        messages.error(request, f'Error al crear la reserva: {str(e)}')
        errors.append(f"Error al guardar la reserva: {str(e)}")
        return render(request, 'register.html', {'errors': errors, 'rooms': rooms})

@staff_member_required
def admin(request):
    """Panel administrativo avanzado con filtros y métricas"""
    qs = Reserva.objects.select_related('configuracion_salon','configuracion_salon__salon').all()

    # Cambio de estado rápido via POST
    if request.method == 'POST' and request.user.is_staff:
        reserva_id = request.POST.get('reserva_id')
        nuevo_estado = request.POST.get('nuevo_estado')
        if reserva_id and nuevo_estado in ['PENDIENTE','CONFIRMADA','CANCELADA','COMPLETADA']:
            try:
                # Check permissions: assistants may confirm/reject if they have the custom permission or are in the group
                from .utils import is_admin_general, is_asistente

                r = qs.get(id=reserva_id)

                # Allow if user is admin general
                if is_admin_general(request.user):
                    allowed = True
                else:
                    # Assistant must have explicit permiso can_confirm_reserva/can_reject_reserva (or change_reserva)
                    if nuevo_estado == 'CONFIRMADA':
                        allowed = request.user.has_perm('reservas.can_confirm_reserva') or request.user.has_perm('reservas.change_reserva')
                    elif nuevo_estado == 'CANCELADA':
                        allowed = request.user.has_perm('reservas.can_reject_reserva') or request.user.has_perm('reservas.change_reserva')
                    else:
                        # allow marking pending/completed if user has generic change permission
                        allowed = request.user.has_perm('reservas.change_reserva')

                if not allowed:
                    messages.error(request, 'No tienes permiso para cambiar el estado de la reserva.')
                else:
                    estado_anterior = r.estado
                    r.estado = nuevo_estado
                    r.save()
                    messages.success(request, f'Reserva #{r.id} cambió de {estado_anterior} a {nuevo_estado}')
            except Reserva.DoesNotExist:
                messages.error(request, 'Reserva no encontrada')

    # Filtros GET
    estado = request.GET.get('estado')
    tipo = request.GET.get('tipo')  # SOCIO / PARTICULAR
    salon = request.GET.get('salon')  # nombre salon
    desde = request.GET.get('desde')  # yyyy-mm-dd
    hasta = request.GET.get('hasta')

    if estado:
        qs = qs.filter(estado=estado)
    if tipo:
        qs = qs.filter(tipo_cliente=tipo)
    if salon:
        qs = qs.filter(configuracion_salon__salon__nombre__icontains=salon)
    if desde:
        qs = qs.filter(fecha_evento__gte=desde)
    if hasta:
        qs = qs.filter(fecha_evento__lte=hasta)

    qs = qs.order_by('-fecha_creacion')

    # Métricas
    total_reservas = qs.count()
    pendientes = qs.filter(estado='PENDIENTE').count()
    confirmadas = qs.filter(estado='CONFIRMADA').count()
    canceladas = qs.filter(estado='CANCELADA').count()
    completadas = qs.filter(estado='COMPLETADA').count()

    # Ingresos estimados por estado (solo confirmadas + completadas)
    ingresos = qs.filter(estado__in=['CONFIRMADA','COMPLETADA']).aggregate(total=Sum('precio_total'))['total'] or 0

    # Distribución por salón
    salas_stats = (
        qs.values('configuracion_salon__salon__nombre')
          .annotate(cantidad=Count('id'), monto=Sum('precio_total'))
          .order_by('-cantidad')[:10]
    )

    context = {
        'reservas': qs,
        'total_reservas': total_reservas,
        'pendientes': pendientes,
        'confirmadas': confirmadas,
        'canceladas': canceladas,
        'completadas': completadas,
        'ingresos': ingresos,
        'salas_stats': salas_stats,
        'estados': ['PENDIENTE','CONFIRMADA','CANCELADA','COMPLETADA'],
        'f_estado': estado or '',
        'f_tipo': tipo or '',
        'f_salon': salon or '',
        'f_desde': desde or '',
        'f_hasta': hasta or '',
    }

    return render(request, 'admin.html', context)

@login_required
def logout_view(request):
    """Logout seguro solo vía POST para evitar 405 en algunos navegadores y extensiones."""
    if request.method == 'POST':
        logout(request)
        messages.info(request, 'Sesión cerrada correctamente')
    return redirect('reservas:index')

@staff_member_required
def borrar_reservas(request):
    """Borrar todas las reservas (solo staff)"""
    if request.method == 'POST':
        count = Reserva.objects.count()
        Reserva.objects.all().delete()
        messages.warning(request, f'Se eliminaron {count} reserva(s) correctamente')
        return redirect('reservas:panel')
    return redirect('reservas:panel')

@staff_member_required
def borrar_reserva_individual(request, reserva_id):
    """Borrar una reserva específica (solo staff)"""
    if request.method == 'POST':
        try:
            reserva = get_object_or_404(Reserva, id=reserva_id)
            cliente = reserva.nombre_cliente
            reserva.delete()
            messages.success(request, f'Reserva #{reserva_id} de {cliente} eliminada correctamente')
        except Exception as e:
            messages.error(request, f'Error al eliminar la reserva: {str(e)}')
    return redirect('reservas:panel')

@staff_member_required
def export_reservas_csv(request):
    """Exportar reservas a Excel con análisis de ganancias"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.db.models import Sum, Count, Q
        from django.db.models.functions import TruncMonth
    except ImportError:
        # Fallback a CSV si openpyxl no está instalado
        return export_reservas_csv_fallback(request)
    
    qs = Reserva.objects.select_related('configuracion_salon','configuracion_salon__salon').all()
    
    # Aplicar filtros GET
    estado = request.GET.get('estado')
    tipo = request.GET.get('tipo')
    salon = request.GET.get('salon')
    desde = request.GET.get('desde')
    hasta = request.GET.get('hasta')
    
    if estado:
        qs = qs.filter(estado=estado)
    if tipo:
        qs = qs.filter(tipo_cliente=tipo)
    if salon:
        qs = qs.filter(configuracion_salon__salon__nombre__icontains=salon)
    if desde:
        qs = qs.filter(fecha_evento__gte=desde)
    if hasta:
        qs = qs.filter(fecha_evento__lte=hasta)
    
    # Crear workbook
    wb = openpyxl.Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    total_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # HOJA 1: Listado de Reservas
    ws1 = wb.active
    ws1.title = "Reservas"
    
    headers = ['ID', 'Cliente', 'Email', 'Teléfono', 'Entidad/Empresa', 'Salón', 'Configuración', 
               'Fecha Evento', 'Hora Inicio', 'Duración', 'Personas', 'Tiempo Decoración', 
               'Tipo Cliente', 'Precio Salón', 'Servicios Adicionales', 'Subtotal Servicios', 
               'Precio Total', 'Estado', 'Fecha Reserva', 'Observaciones']
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Additional styles: alternate row fill and estado colors
    alt_fill = PatternFill(start_color="F7FAFF", end_color="F7FAFF", fill_type="solid")
    estado_fills = {
        'CONFIRMADA': PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid"),
        'COMPLETADA': PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid"),
        'PENDIENTE': PatternFill(start_color="FFF3BF", end_color="FFF3BF", fill_type="solid"),
        'CANCELADA': PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    }

    row = 2
    total_ingresos = 0
    for r in qs:
        ws1.cell(row, 1, r.id)
        ws1.cell(row, 2, r.nombre_cliente)
        ws1.cell(row, 3, r.email_cliente)
        ws1.cell(row, 4, r.telefono_cliente)
        ws1.cell(row, 5, r.nombre_entidad or 'N/A')
        ws1.cell(row, 6, r.configuracion_salon.salon.nombre)
        ws1.cell(row, 7, r.configuracion_salon.get_tipo_configuracion_display())
        ws1.cell(row, 8, r.fecha_evento.strftime('%Y-%m-%d'))
        ws1.cell(row, 9, r.hora_inicio.strftime('%H:%M') if r.hora_inicio else 'N/A')
        ws1.cell(row, 10, r.get_duracion_display())
        ws1.cell(row, 11, r.numero_personas)
        ws1.cell(row, 12, f"{r.tiempo_decoracion}h")
        ws1.cell(row, 13, r.get_tipo_cliente_display())
        ws1.cell(row, 14, float(r.precio_total))
        
        # Servicios Adicionales
        servicios_adicionales = r.servicios_adicionales.all()
        if servicios_adicionales:
            servicios_detalle = []
            subtotal_servicios = 0
            for sa in servicios_adicionales:
                servicios_detalle.append(f"{sa.servicio.nombre} x{sa.cantidad} = ${sa.subtotal:,.0f}")
                subtotal_servicios += float(sa.subtotal)
            ws1.cell(row, 15, "\n".join(servicios_detalle))
            ws1.cell(row, 16, subtotal_servicios)
            ws1.cell(row, 16).number_format = '$#,##0.00'
        else:
            ws1.cell(row, 15, "N/A")
            ws1.cell(row, 16, 0)
        
        ws1.cell(row, 17, float(r.precio_total))
        ws1.cell(row, 18, r.get_estado_display())
        ws1.cell(row, 19, r.fecha_creacion.strftime('%Y-%m-%d %H:%M'))
        ws1.cell(row, 20, r.observaciones or '')
        
        if r.estado in ['CONFIRMADA', 'COMPLETADA']:
            total_ingresos += float(r.precio_total)
        
        for col in range(1, 21):
            ws1.cell(row, col).border = border
            # Wrap largo texto de servicios / observaciones
            if col == 15 or col == 20:
                ws1.cell(row, col).alignment = Alignment(wrap_text=True)

        # Apply alternate shading for readabilty
        if row % 2 == 0:
            for col in range(1, 21):
                ws1.cell(row, col).fill = alt_fill

        # Color by estado (column Q / 18)
        try:
            estado_cell = ws1.cell(row, 18)
            estado_key = r.estado
            if estado_key in estado_fills:
                estado_cell.fill = estado_fills[estado_key]
        except Exception:
            pass
        
        # Ajustar altura de fila si hay servicios
        if servicios_adicionales and len(servicios_adicionales) > 1:
            ws1.row_dimensions[row].height = 15 * len(servicios_adicionales)
        
        row += 1
    
    # Total al final
    ws1.cell(row, 16, "TOTAL INGRESOS:").font = total_font
    ws1.cell(row, 17, total_ingresos).font = total_font
    ws1.cell(row, 17).number_format = '$#,##0.00'
    ws1.cell(row, 16).fill = total_fill
    ws1.cell(row, 17).fill = total_fill

    # Freeze header and enable autofilter for easier navigation
    ws1.freeze_panes = ws1['A2']
    last_row = row
    ws1.auto_filter.ref = f"A1:T{last_row}"
    
    # Ajustar anchos
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 30
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 25
    ws1.column_dimensions['F'].width = 20
    ws1.column_dimensions['G'].width = 15
    ws1.column_dimensions['H'].width = 15
    ws1.column_dimensions['I'].width = 12
    ws1.column_dimensions['J'].width = 12
    ws1.column_dimensions['K'].width = 10
    ws1.column_dimensions['L'].width = 12
    ws1.column_dimensions['M'].width = 15
    ws1.column_dimensions['N'].width = 15
    ws1.column_dimensions['O'].width = 40  # Servicios Adicionales
    ws1.column_dimensions['P'].width = 18  # Subtotal Servicios
    ws1.column_dimensions['Q'].width = 15  # Precio Total
    ws1.column_dimensions['R'].width = 12  # Estado
    ws1.column_dimensions['S'].width = 18  # Fecha Reserva
    ws1.column_dimensions['T'].width = 30  # Observaciones
    
    # HOJA 2: Resumen por Mes
    ws2 = wb.create_sheet("Resumen Mensual")
    
    # Headers
    ws2.cell(1, 1, "Mes").fill = header_fill
    ws2.cell(1, 1).font = header_font
    ws2.cell(1, 2, "Total Reservas").fill = header_fill
    ws2.cell(1, 2).font = header_font
    ws2.cell(1, 3, "Confirmadas").fill = header_fill
    ws2.cell(1, 3).font = header_font
    ws2.cell(1, 4, "Ingresos").fill = header_fill
    ws2.cell(1, 4).font = header_font
    
    # Agrupar por mes
    reservas_por_mes = Reserva.objects.filter(
        Q(estado='CONFIRMADA') | Q(estado='COMPLETADA')
    ).annotate(
        mes=TruncMonth('fecha_evento')
    ).values('mes').annotate(
        total=Count('id'),
        ingresos=Sum('precio_total')
    ).order_by('mes')
    
    row = 2
    total_mes = 0
    for item in reservas_por_mes:
        ws2.cell(row, 1, item['mes'].strftime('%B %Y'))
        ws2.cell(row, 2, item['total'])
        ws2.cell(row, 3, item['total'])
        ws2.cell(row, 4, float(item['ingresos'] or 0))
        ws2.cell(row, 4).number_format = '$#,##0.00'
        total_mes += float(item['ingresos'] or 0)
        row += 1
    
    # Total anual
    ws2.cell(row, 3, "TOTAL AÑO:").font = total_font
    ws2.cell(row, 4, total_mes).font = total_font
    ws2.cell(row, 4).number_format = '$#,##0.00'
    ws2.cell(row, 3).fill = total_fill
    ws2.cell(row, 4).fill = total_fill
    
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 20
    
    # HOJA 3: Resumen por Salón
    ws3 = wb.create_sheet("Resumen por Salón")
    
    ws3.cell(1, 1, "Salón").fill = header_fill
    ws3.cell(1, 1).font = header_font
    ws3.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
    ws3.cell(1, 2, "Total Reservas").fill = header_fill
    ws3.cell(1, 2).font = header_font
    ws3.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
    ws3.cell(1, 3, "Confirmadas").fill = header_fill
    ws3.cell(1, 3).font = header_font
    ws3.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
    ws3.cell(1, 4, "Ingresos").fill = header_fill
    ws3.cell(1, 4).font = header_font
    ws3.cell(1, 4).alignment = Alignment(horizontal='center', vertical='center')
    ws3.cell(1, 5, "Promedio/Reserva").fill = header_fill
    ws3.cell(1, 5).font = header_font
    ws3.cell(1, 5).alignment = Alignment(horizontal='center', vertical='center')
    
    reservas_por_salon = Reserva.objects.filter(
        Q(estado='CONFIRMADA') | Q(estado='COMPLETADA')
    ).values(
        'configuracion_salon__salon__nombre'
    ).annotate(
        total=Count('id'),
        ingresos=Sum('precio_total')
    ).order_by('-ingresos')
    
    row = 2
    total_salon_ingresos = 0
    for item in reservas_por_salon:
        salon_nombre = item['configuracion_salon__salon__nombre']
        total_reservas = item['total']
        ingresos = float(item['ingresos'] or 0)
        promedio = ingresos / total_reservas if total_reservas > 0 else 0
        
        ws3.cell(row, 1, salon_nombre)
        ws3.cell(row, 2, total_reservas).alignment = Alignment(horizontal='center')
        ws3.cell(row, 3, total_reservas).alignment = Alignment(horizontal='center')
        ws3.cell(row, 4, ingresos)
        ws3.cell(row, 4).number_format = '$#,##0.00'
        ws3.cell(row, 5, promedio)
        ws3.cell(row, 5).number_format = '$#,##0.00'
        
        total_salon_ingresos += ingresos
        
        for col in range(1, 6):
            ws3.cell(row, col).border = border
        
        row += 1
    
    # Total
    ws3.cell(row, 3, "TOTAL:").font = total_font
    ws3.cell(row, 4, total_salon_ingresos).font = total_font
    ws3.cell(row, 4).number_format = '$#,##0.00'
    ws3.cell(row, 3).fill = total_fill
    ws3.cell(row, 4).fill = total_fill
    
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 15
    ws3.column_dimensions['C'].width = 15
    ws3.column_dimensions['D'].width = 20
    ws3.column_dimensions['E'].width = 20
    
    # HOJA 4: Dashboard Anual - Gestión Empresarial
    ws4 = wb.create_sheet("Dashboard Anual")
    
    # Título principal
    ws4.merge_cells('A1:F1')
    titulo = ws4.cell(1, 1, "CLUB EL META - ANÁLISIS ANUAL DE GESTIÓN")
    titulo.font = Font(bold=True, size=16, color="FFFFFF")
    titulo.fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    titulo.alignment = Alignment(horizontal='center', vertical='center')
    
    # Fecha del reporte
    ws4.merge_cells('A2:F2')
    fecha_reporte = ws4.cell(2, 1, f"Reporte generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    fecha_reporte.alignment = Alignment(horizontal='center')
    fecha_reporte.font = Font(italic=True, size=10)
    
    # Métricas generales
    from django.db.models.functions import TruncYear
    
    ws4.cell(4, 1, "MÉTRICAS GENERALES").font = Font(bold=True, size=13)
    ws4.cell(4, 1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws4.merge_cells('A4:F4')
    
    total_reservas_año = Reserva.objects.filter(Q(estado='CONFIRMADA') | Q(estado='COMPLETADA')).count()
    total_ingresos_año = float(Reserva.objects.filter(Q(estado='CONFIRMADA') | Q(estado='COMPLETADA')).aggregate(Sum('precio_total'))['precio_total__sum'] or 0)
    promedio_reserva = total_ingresos_año / total_reservas_año if total_reservas_año > 0 else 0
    
    metricas = [
        ["Total Reservas Confirmadas:", total_reservas_año, ""],
        ["Total Ingresos:", f"${total_ingresos_año:,.2f}", ""],
        ["Promedio por Reserva:", f"${promedio_reserva:,.2f}", ""],
    ]
    
    row = 5
    for metrica in metricas:
        ws4.cell(row, 1, metrica[0]).font = Font(bold=True)
        ws4.cell(row, 2, metrica[1]).font = Font(size=12)
        ws4.cell(row, 2).alignment = Alignment(horizontal='left')
        row += 1
    
    # Análisis por Año
    ws4.cell(row + 1, 1, "INGRESOS POR AÑO").font = Font(bold=True, size=13)
    ws4.cell(row + 1, 1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws4.merge_cells(f'A{row+1}:F{row+1}')
    
    row += 2
    ws4.cell(row, 1, "Año").fill = header_fill
    ws4.cell(row, 1).font = header_font
    ws4.cell(row, 2, "Total Reservas").fill = header_fill
    ws4.cell(row, 2).font = header_font
    ws4.cell(row, 3, "Ingresos").fill = header_fill
    ws4.cell(row, 3).font = header_font
    ws4.cell(row, 4, "Promedio/Reserva").fill = header_fill
    ws4.cell(row, 4).font = header_font
    
    reservas_por_año = Reserva.objects.filter(
        Q(estado='CONFIRMADA') | Q(estado='COMPLETADA')
    ).annotate(
        año=TruncYear('fecha_evento')
    ).values('año').annotate(
        total=Count('id'),
        ingresos=Sum('precio_total')
    ).order_by('año')
    
    row += 1
    for item in reservas_por_año:
        año_str = item['año'].strftime('%Y')
        total_res = item['total']
        ingresos_año = float(item['ingresos'] or 0)
        promedio = ingresos_año / total_res if total_res > 0 else 0
        
        ws4.cell(row, 1, año_str).alignment = Alignment(horizontal='center')
        ws4.cell(row, 2, total_res).alignment = Alignment(horizontal='center')
        ws4.cell(row, 3, ingresos_año)
        ws4.cell(row, 3).number_format = '$#,##0.00'
        ws4.cell(row, 4, promedio)
        ws4.cell(row, 4).number_format = '$#,##0.00'
        
        for col in range(1, 5):
            ws4.cell(row, col).border = border
        
        row += 1
    
    # Análisis por Tipo de Cliente
    ws4.cell(row + 1, 1, "ANÁLISIS POR TIPO DE CLIENTE").font = Font(bold=True, size=13)
    ws4.cell(row + 1, 1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws4.merge_cells(f'A{row+1}:F{row+1}')
    
    row += 2
    ws4.cell(row, 1, "Tipo Cliente").fill = header_fill
    ws4.cell(row, 1).font = header_font
    ws4.cell(row, 2, "Cantidad").fill = header_fill
    ws4.cell(row, 2).font = header_font
    ws4.cell(row, 3, "Ingresos").fill = header_fill
    ws4.cell(row, 3).font = header_font
    ws4.cell(row, 4, "% del Total").fill = header_fill
    ws4.cell(row, 4).font = header_font
    
    por_tipo = Reserva.objects.filter(
        Q(estado='CONFIRMADA') | Q(estado='COMPLETADA')
    ).values('tipo_cliente').annotate(
        total=Count('id'),
        ingresos=Sum('precio_total')
    ).order_by('-ingresos')
    
    row += 1
    for item in por_tipo:
        tipo_nombre = "Socio" if item['tipo_cliente'] == 'SOCIO' else "Particular"
        cantidad = item['total']
        ingresos_tipo = float(item['ingresos'] or 0)
        porcentaje = (ingresos_tipo / float(total_ingresos_año) * 100) if total_ingresos_año > 0 else 0
        
        ws4.cell(row, 1, tipo_nombre)
        ws4.cell(row, 2, cantidad).alignment = Alignment(horizontal='center')
        ws4.cell(row, 3, ingresos_tipo)
        ws4.cell(row, 3).number_format = '$#,##0.00'
        ws4.cell(row, 4, f"{porcentaje:.1f}%").alignment = Alignment(horizontal='center')
        
        for col in range(1, 5):
            ws4.cell(row, col).border = border
        
        row += 1
    
    # Ajustar anchos Dashboard
    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['C'].width = 20
    ws4.column_dimensions['D'].width = 20
    ws4.column_dimensions['E'].width = 15
    ws4.column_dimensions['F'].width = 15
    
    # Guardar en respuesta
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reservas_club_el_meta_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response


def export_reservas_csv_fallback(request):
    """Fallback CSV si openpyxl no está disponible"""
    qs = Reserva.objects.select_related('configuracion_salon','configuracion_salon__salon').all()
    
    estado = request.GET.get('estado')
    tipo = request.GET.get('tipo')
    salon = request.GET.get('salon')
    desde = request.GET.get('desde')
    hasta = request.GET.get('hasta')
    
    if estado:
        qs = qs.filter(estado=estado)
    if tipo:
        qs = qs.filter(tipo_cliente=tipo)
    if salon:
        qs = qs.filter(configuracion_salon__salon__nombre__icontains=salon)
    if desde:
        qs = qs.filter(fecha_evento__gte=desde)
    if hasta:
        qs = qs.filter(fecha_evento__lte=hasta)
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="reservas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    response.write('\ufeff')

    # Use semicolon delimiter for better Excel compatibility in locales that use comma as decimal
    writer = csv.writer(response, delimiter=';')

    def fmt_money(value):
        try:
            n = float(value or 0)
        except Exception:
            n = 0.0
        s = f"{n:,.2f}"
        # Convert 1,234,567.89 -> 1.234.567,89 (Spanish format)
        s = s.replace(',', 'X').replace('.', ',').replace('X', '.')
        return s

    # Title
    writer.writerow([f"Listado de Reservas - Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    writer.writerow([])

    # Header (match spreadsheet columns)
    headers = ['ID', 'Cliente', 'Email', 'Teléfono', 'Entidad/Empresa', 'Salón', 'Configuración',
               'Fecha Evento', 'Hora Inicio', 'Duración', 'Personas', 'Tiempo Decoración', 'Tipo Cliente',
               'Precio Salón', 'Servicios Adicionales', 'Subtotal Servicios', 'Precio Total', 'Estado', 'Fecha Reserva', 'Observaciones']
    writer.writerow(headers)

    total_ingresos = 0
    for r in qs:
        # Servicios adicionales
        servicios_adicionales = []
        subtotal_servicios = 0
        try:
            servicios_qs = r.servicios_adicionales.all()
            for sa in servicios_qs:
                servicios_adicionales.append(f"{sa.servicio.nombre} x{sa.cantidad}")
                subtotal_servicios += float(sa.subtotal or 0)
        except Exception:
            servicios_qs = []

        precio_total = float(r.precio_total or 0)
        # Estimar precio salón (precio total menos servicios si aplica)
        precio_salon = precio_total - subtotal_servicios if precio_total >= subtotal_servicios else precio_total

        writer.writerow([
            r.id,
            r.nombre_cliente,
            r.email_cliente,
            r.telefono_cliente,
            r.nombre_entidad or 'N/A',
            r.configuracion_salon.salon.nombre if r.configuracion_salon and r.configuracion_salon.salon else 'N/A',
            r.configuracion_salon.get_tipo_configuracion_display() if r.configuracion_salon else 'N/A',
            r.fecha_evento.strftime('%Y-%m-%d') if r.fecha_evento else '',
            r.hora_inicio.strftime('%H:%M') if getattr(r, 'hora_inicio', None) else '',
            r.get_duracion_display(),
            r.numero_personas,
            f"{r.tiempo_decoracion}h" if getattr(r, 'tiempo_decoracion', None) is not None else '',
            r.get_tipo_cliente_display(),
            fmt_money(precio_salon),
            " | ".join(servicios_adicionales) if servicios_adicionales else 'N/A',
            fmt_money(subtotal_servicios),
            fmt_money(precio_total),
            r.get_estado_display(),
            r.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S') if r.fecha_creacion else '',
            (r.observaciones or '').replace('\n', ' ')
        ])

        if r.estado in ['CONFIRMADA', 'COMPLETADA']:
            total_ingresos += precio_total

    # Separator and totals
    writer.writerow([])
    writer.writerow(['', '', '', '', '', '', '', '', '', '', '', '', '', 'TOTAL INGRESOS:', fmt_money(total_ingresos)])

    # Conteo por Estado
    writer.writerow([])
    writer.writerow(['Conteo por Estado'])
    estados = ['PENDIENTE', 'CONFIRMADA', 'CANCELADA', 'COMPLETADA']
    for e in estados:
        writer.writerow([e, qs.filter(estado=e).count()])

    # Por salón
    writer.writerow([])
    writer.writerow(['Por Salón', 'Cantidad Reservas', 'Monto Total'])
    salas = qs.values('configuracion_salon__salon__nombre').annotate(cantidad=Count('id'), monto=Sum('precio_total')).order_by('-cantidad')
    for s in salas:
        nombre = s.get('configuracion_salon__salon__nombre') or 'Sin salón'
        cantidad = s.get('cantidad') or 0
        monto = float(s.get('monto') or 0)
        writer.writerow([nombre, cantidad, fmt_money(monto)])

    # Servicios adicionales totales
    try:
        from .models import ReservaServicioAdicional
        servicios_totales = ReservaServicioAdicional.objects.filter(reserva__in=qs)
        ingresos_servicios = sum([float(s.subtotal or 0) for s in servicios_totales])
        writer.writerow([])
        writer.writerow(['Ingresos por Servicios Adicionales:', fmt_money(ingresos_servicios)])
    except Exception:
        pass

    return response


def check_availability(request):
    """API endpoint para verificar disponibilidad de espacios en una fecha específica"""
    fecha_str = request.GET.get('fecha', '')
    
    try:
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except:
        return JsonResponse({'error': 'Fecha inválida'}, status=400)
    
    # Obtener todas las configuraciones
    configuraciones = ConfiguracionSalon.objects.filter(salon__disponible=True).select_related('salon')
    
    # Verificar cuáles salones están bloqueados
    salones_bloqueados = set()
    bloqueos = BloqueoEspacio.objects.filter(
        activo=True,
        fecha_inicio__lte=fecha_obj,
        fecha_fin__gte=fecha_obj
    ).select_related('salon')
    
    bloqueos_info = {}
    for bloqueo in bloqueos:
        salones_bloqueados.add(bloqueo.salon.id)
        bloqueos_info[bloqueo.salon.id] = {
            'motivo': bloqueo.get_motivo_display(),
            'descripcion': bloqueo.descripcion
        }
    
    # Construir respuesta
    result = []
    for config in configuraciones:
        esta_bloqueado = config.salon.id in salones_bloqueados
        item = {
            'id': config.id,
            'salon_id': config.salon.id,
            'salon_nombre': config.salon.nombre,
            'configuracion': config.get_tipo_configuracion_display(),
            'capacidad': config.capacidad,
            'disponible': not esta_bloqueado,
        }
        
        if esta_bloqueado:
            item['motivo_bloqueo'] = bloqueos_info[config.salon.id]['motivo']
            item['descripcion_bloqueo'] = bloqueos_info[config.salon.id]['descripcion']
        
        result.append(item)
    
    return JsonResponse({'espacios': result, 'fecha': fecha_str})


def get_bloqueos_salon(request):
    """API endpoint para obtener todos los bloqueos de un salón específico"""
    salon_id = request.GET.get('salon_id', '')
    
    if not salon_id:
        return JsonResponse({'error': 'salon_id requerido'}, status=400)
    
    try:
        from datetime import date
        hoy = date.today()
        
        # Obtener bloqueos activos del salón que aún no han terminado
        # Algunos registros en BD pueden tener `fecha_fin` anterior a `fecha_inicio` (datos inconsistentes).
        # Para evitar omitir bloqueos futuros en el formulario (aunque fecha_fin esté mal),
        # incluimos bloqueos cuya fecha_fin sea >= hoy OR cuya fecha_inicio sea >= hoy.
        from django.db.models import Q
        bloqueos = BloqueoEspacio.objects.filter(
            Q(salon_id=salon_id),
            Q(activo=True),
            (Q(fecha_fin__gte=hoy) | Q(fecha_inicio__gte=hoy))
        ).order_by('fecha_inicio')
        
        result = []
        for bloqueo in bloqueos:
            # Normalize dates: if fecha_fin is before fecha_inicio, treat fecha_fin as fecha_inicio
            fecha_inicio = bloqueo.fecha_inicio
            fecha_fin = bloqueo.fecha_fin
            if fecha_fin < fecha_inicio:
                fecha_fin = fecha_inicio

            result.append({
                'fecha_inicio': fecha_inicio.strftime('%Y-%m-%d'),
                'fecha_fin': fecha_fin.strftime('%Y-%m-%d'),
                'motivo': bloqueo.get_motivo_display(),
                'descripcion': bloqueo.descripcion
            })
        
        return JsonResponse({'bloqueos': result})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

# Vista para el calendario
@login_required
def calendario(request):
    """Vista principal del calendario"""
    espacios = Salon.objects.all().order_by('nombre')
    return render(request, 'calendario.html', {
        'espacios': espacios
    })


@staff_member_required
def reportes_dashboard(request):
    """Dashboard de reportes: calcula métricas y series para gráficas.

    - Filtra por rango de fechas y por espacio (opcional)
    - Devuelve: total reservas, ingresos, promedio de personas, ingresos por día y ranking por salón
    """
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para acceder a los reportes.')
        return redirect('reservas:index')

    from django.db.models import Sum, Count
    from datetime import timedelta
    # Parámetros
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    espacio_id = request.GET.get('espacio_id')
    preset = request.GET.get('preset')  # '30d', 'month', 'year', 'custom'

    # Por defecto: últimos 30 días (hasta 30 días en el futuro para incluir reservas próximas)
    from datetime import timedelta
    hoy = datetime.now().date()

    # If a preset is provided and it's not 'custom', compute fecha_inicio/fecha_fin from it
    if preset and preset != 'custom':
        try:
            import calendar
            def subtract_months(d, months):
                m = d.month - months
                y = d.year
                while m <= 0:
                    m += 12
                    y -= 1
                # adjust day to last day of target month if needed
                last_day = calendar.monthrange(y, m)[1]
                day = min(d.day, last_day)
                return d.replace(year=y, month=m, day=day)

            if preset == '7d':
                fecha_fin_obj = hoy
                fecha_inicio_obj = fecha_fin_obj - timedelta(days=7)
            elif preset == '15d':
                fecha_fin_obj = hoy
                fecha_inicio_obj = fecha_fin_obj - timedelta(days=15)
            elif preset == '30d':
                fecha_fin_obj = hoy
                fecha_inicio_obj = fecha_fin_obj - timedelta(days=30)
            elif preset == '90d':
                fecha_fin_obj = hoy
                fecha_inicio_obj = fecha_fin_obj - timedelta(days=90)
            elif preset == '6m':
                fecha_fin_obj = hoy
                fecha_inicio_obj = subtract_months(fecha_fin_obj, 6)
            elif preset == 'month':
                # Mes actual: incluir hasta el último día del mes (posibles reservas futuras)
                first_day = hoy.replace(day=1)
                last_day_num = calendar.monthrange(hoy.year, hoy.month)[1]
                last_day = hoy.replace(day=last_day_num)
                fecha_inicio_obj = first_day
                fecha_fin_obj = last_day
            elif preset == 'year':
                fecha_inicio_obj = hoy.replace(month=1, day=1)
                fecha_fin_obj = hoy.replace(month=12, day=31)
            else:
                # fallback to default 30d
                fecha_fin_obj = hoy + timedelta(days=30)
                fecha_inicio_obj = fecha_fin_obj - timedelta(days=30)
        except Exception:
            fecha_fin_obj = hoy + timedelta(days=30)
            fecha_inicio_obj = fecha_fin_obj - timedelta(days=30)
    else:
        # No preset or custom preset: use explicit fecha_inicio/fecha_fin if provided
        if not fecha_fin:
            # Por defecto termina en hoy para que presets sean consistentes
            fecha_fin_obj = hoy
        else:
            try:
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            except:
                fecha_fin_obj = hoy

        if not fecha_inicio:
            fecha_inicio_obj = fecha_fin_obj - timedelta(days=30)
        else:
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            except:
                fecha_inicio_obj = fecha_fin_obj - timedelta(days=30)

    # Query base: reservas confirmadas o completadas en el rango
    qs = Reserva.objects.filter(
        estado__in=['CONFIRMADA', 'COMPLETADA'],
        fecha_evento__gte=fecha_inicio_obj,
        fecha_evento__lte=fecha_fin_obj
    ).select_related('configuracion_salon__salon')

    if espacio_id:
        try:
            qs = qs.filter(configuracion_salon__salon_id=int(espacio_id))
        except ValueError:
            pass

    # Métricas
    total_reservas = qs.count()
    ingresos_agg = qs.aggregate(total=Sum('precio_total'))['total'] or 0
    ingresos_totales = float(ingresos_agg)

    personas_agg = qs.aggregate(total_personas=Sum('numero_personas'))['total_personas'] or 0
    promedio_personas = round(float(personas_agg) / total_reservas, 1) if total_reservas > 0 else 0

    # Total acumulado (sin filtrar por fecha) para comparar con el panel
    try:
        total_all = Reserva.objects.filter(estado__in=['CONFIRMADA', 'COMPLETADA']).aggregate(total=Sum('precio_total'))['total'] or 0
        ingresos_totales_all = float(total_all)
    except Exception:
        ingresos_totales_all = ingresos_totales

    # Ingresos por día: agregación por fecha_evento
    reservas_por_dia = qs.values('fecha_evento').annotate(total=Sum('precio_total'), count=Count('id')).order_by('fecha_evento')
    ingresos_por_dia_map = {item['fecha_evento'].strftime('%Y-%m-%d'): float(item['total'] or 0) for item in reservas_por_dia}

    # Preparar rango completo de fechas
    current = fecha_inicio_obj
    fechas = []
    ingresos_values = []
    while current <= fecha_fin_obj:
        key = current.strftime('%Y-%m-%d')
        fechas.append(key)
        ingresos_values.append(ingresos_por_dia_map.get(key, 0))
        current += timedelta(days=1)

    # Ranking de espacios por ingresos
    espacios_qs = qs.values('configuracion_salon__salon__nombre').annotate(monto=Sum('precio_total'), cantidad=Count('id')).order_by('-monto')
    espacios_labels = [item['configuracion_salon__salon__nombre'] for item in espacios_qs]
    espacios_values = [float(item['monto'] or 0) for item in espacios_qs]

    # Servicios adicionales (si existe el modelo)
    ingresos_servicios = 0
    try:
        from .models import ReservaServicioAdicional
        servicios_totales = ReservaServicioAdicional.objects.filter(reserva__in=qs)
        ingresos_servicios = sum([float(s.subtotal or 0) for s in servicios_totales])
    except Exception:
        ingresos_servicios = 0

    espacios = Salon.objects.all().order_by('nombre')

    context = {
        'fecha_inicio': fecha_inicio_obj.strftime('%Y-%m-%d'),
        'fecha_fin': fecha_fin_obj.strftime('%Y-%m-%d'),
        'espacio_id': espacio_id,
        'espacios': espacios,
        'total_reservas': total_reservas,
        'ingresos_totales': ingresos_totales,
        'ingresos_totales_all': ingresos_totales_all,
        'ingresos_servicios': ingresos_servicios,
        'ingresos_gran_total': ingresos_totales + ingresos_servicios,
        'promedio_personas': promedio_personas,
        'fechas_labels': fechas,
        'ingresos_values': ingresos_values,
        'espacios_labels': espacios_labels,
        'espacios_values': espacios_values,
        'preset': preset or '30d',
    }

    return render(request, 'reportes.html', context)

# API para obtener eventos del calendario
@login_required
def get_calendar_events(request):
    """Obtiene todos los eventos (reservas y bloqueos) para el calendario"""
    try:
        espacio_id = request.GET.get('espacio_id', None)
        events = []
        
        # Filtrar reservas
        reservas = Reserva.objects.select_related('configuracion_salon__salon')
        if espacio_id:
            reservas = reservas.filter(configuracion_salon__salon_id=espacio_id)
        
        for reserva in reservas:
            try:
                # Formatear hora de inicio - manejar diferentes tipos
                if hasattr(reserva.hora_inicio, 'strftime'):
                    hora_str = reserva.hora_inicio.strftime('%H:%M:%S')
                else:
                    # Si es string, asegurar formato correcto
                    hora_str = str(reserva.hora_inicio)
                    if len(hora_str) == 5:  # HH:MM
                        hora_str = hora_str + ':00'
                
                espacio_nombre = reserva.configuracion_salon.salon.nombre
                config_nombre = reserva.configuracion_salon.tipo_configuracion
                duracion_num = 4 if reserva.duracion == '4H' else 8
                
                # Calcular hora de fin para mejor visualización
                from datetime import datetime as dt_calc, timedelta
                hora_inicio_dt = dt_calc.strptime(hora_str, '%H:%M:%S')
                hora_fin_dt = hora_inicio_dt + timedelta(hours=duracion_num)
                
                events.append({
                    'id': f'reserva_{reserva.id}',
                    'title': f'{reserva.nombre_cliente} - {espacio_nombre} ({config_nombre})',
                    'start': f'{reserva.fecha_evento}T{hora_str}',
                    'end': f'{reserva.fecha_evento}T{hora_fin_dt.strftime("%H:%M:%S")}',
                    'backgroundColor': '#2563eb',
                    'borderColor': '#1e40af',
                    'textColor': '#ffffff',
                    'extendedProps': {
                        'tipo': 'reserva',
                        'reserva_id': reserva.id,
                        'cliente': reserva.nombre_cliente,
                        'espacio': espacio_nombre,
                        'configuracion': config_nombre,
                        'fecha_evento': reserva.fecha_evento.strftime('%Y-%m-%d'),
                        'hora_inicio': hora_str[:5],  # Solo HH:MM
                        'duracion': duracion_num,
                        'num_personas': reserva.numero_personas,
                        'telefono': reserva.telefono_cliente,
                        'email': reserva.email_cliente,
                        'precio_total': f'{reserva.precio_total:,.0f}'.replace(',', '.')
                    }
                })
            except Exception as e:
                print(f"Error procesando reserva {reserva.id}: {str(e)}")
                continue
        
        # Filtrar bloqueos
        bloqueos = BloqueoEspacio.objects.select_related('salon')
        if espacio_id:
            bloqueos = bloqueos.filter(salon_id=espacio_id)
        
        for bloqueo in bloqueos:
            try:
                events.append({
                    'id': f'bloqueo_{bloqueo.id}',
                    'title': f'BLOQUEADO - {bloqueo.salon.nombre}',
                    'start': bloqueo.fecha_inicio.strftime('%Y-%m-%d'),
                    'end': (bloqueo.fecha_fin + timedelta(days=1)).strftime('%Y-%m-%d'),
                    'backgroundColor': '#dc2626',
                    'borderColor': '#b91c1c',
                    'display': 'background',
                    'extendedProps': {
                        'tipo': 'bloqueo',
                        'espacio': bloqueo.salon.nombre,
                        'fecha_inicio': bloqueo.fecha_inicio.strftime('%Y-%m-%d'),
                        'fecha_fin': bloqueo.fecha_fin.strftime('%Y-%m-%d'),
                        'motivo': bloqueo.get_motivo_display()
                    }
                })
            except Exception as e:
                print(f"Error procesando bloqueo {bloqueo.id}: {str(e)}")
                continue
        
        return JsonResponse(events, safe=False)
    except Exception as e:
        import traceback
        print(f"Error en get_calendar_events: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({'error': str(e), 'traceback': traceback.format_exc()}, status=400)
