from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Sum, Count
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import logout
from django.contrib import messages
from .models import Salon, ConfiguracionSalon, Reserva, BloqueoEspacio, ReservaServicioAdicional
import csv
from datetime import datetime, timedelta, date

# Mapeo de imágenes por salón (hardcoded por ahora)
SALON_IMAGES = {
    'Mi Llanura': [
        "salon mi llanura/milanuraa.avif",
        "salon mi llanura/millanura.avif",
        "salon mi llanura/millanuraaa.avif",
        "salon mi llanura/mika.jpg"
    ],
    'Salón Bar': [
        "salon bar/bar.jpg",
        "salon bar/barsito.jpg"
    ],
    'Salón Empresarial': ["salon empresarial/salonempresarial.jpg", "salon empresarial/salonempresarialxd.jpg"],
    'Terraza': ["salon terraza/terraza.jpg", "salon terraza/1.jpg"],
    'Salón Kiosco': ["salon kiosco/kiosco.jpg", "salon kiosco/kioscoxdd.jpg", "salon kiosco/kioskoxd.jpg"],
    'Salón Presidente': ["salon presidente/presidente.jpg", "salon presidente/presidente1.jpg"]
}

def get_salon_images(nombre_salon):
    """Obtiene las imágenes del salón basado en su nombre"""
    for key in SALON_IMAGES.keys():
        if key in nombre_salon:
            return SALON_IMAGES[key]
    return []

# Validación de código de socio contra base de datos
def is_valid_socio(code: str) -> bool:
    """Valida si el código de socio existe y está activo"""
    if not code:
        return False
    from .models import CodigoSocio
    return CodigoSocio.objects.filter(codigo=code.strip(), activo=True).exists()

def index(request):
    """Página principal - bienvenida"""
    return render(request, 'index.html')

def espacios(request):
    """Página de espacios disponibles - usando datos de la base de datos"""
    from .models import ConfiguracionSalon
    
    salones = Salon.objects.filter(disponible=True).order_by('nombre')
    
    # Convertir salones a formato compatible con template
    # Ahora mostramos UN salón con sus configuraciones como opciones
    rooms = []
    for salon in salones:
        # Obtener todas las configuraciones del salón
        configuraciones = salon.configuraciones.all()
        
        if configuraciones.exists():
            # Usar la primera configuración para mostrar precio/capacidad de referencia
            primera_config = configuraciones.first()
            
            rooms.append({
                'id': salon.id,
                'name': salon.nombre,
                'desc': salon.descripcion,
                # proveer claves explícitas para 4h y 8h que usan los templates/JS
                'price_particular_4h': int(primera_config.precio_particular_4h),
                'price_particular_8h': int(primera_config.precio_particular_8h or primera_config.precio_particular_4h),
                'price_socio_4h': int(primera_config.precio_socio_4h),
                'price_socio_8h': int(primera_config.precio_socio_8h or primera_config.precio_socio_4h),
                # mantener compatibilidad con código previo
                'price_particular': int(primera_config.precio_particular_4h),
                'price_socio': int(primera_config.precio_socio_4h),
                'capacity': primera_config.capacidad,
                'type': 'social',
                'images': get_salon_images(salon.nombre),
                'configuraciones': list(configuraciones)  # Lista de todas las configuraciones
            })
    
    return render(request, 'espacios.html', {'rooms': rooms})

def register(request):
    """Página de registro/reserva"""
    from .models import ConfiguracionSalon
    
    configuraciones = ConfiguracionSalon.objects.filter(salon__disponible=True).select_related('salon').order_by('salon__nombre', 'tipo_configuracion')
    
    # Convertir configuraciones a formato compatible con template
    rooms = []
    for config in configuraciones:
        # Determinar el texto del precio
        if config.precio_socio_4h == 0:
            precio_texto = f"Socio: Cortesía / Particular: ${config.precio_particular_4h:,.0f}"
        else:
            precio_texto = f"Socio: ${config.precio_socio_4h:,.0f} / Particular: ${config.precio_particular_4h:,.0f}"
        
        rooms.append({
            'id': config.id,
            'salon_id': config.salon.id,
            'name': f"{config.salon.nombre} - {config.get_tipo_configuracion_display()}",
            'desc': f"{config.capacidad} personas",
            'price': precio_texto,  # Texto completo del precio
            # claves 4h/8h para JS
            'price_particular_4h': int(config.precio_particular_4h),
            'price_particular_8h': int(config.precio_particular_8h or config.precio_particular_4h),
            'price_socio_4h': int(config.precio_socio_4h),
            'price_socio_8h': int(config.precio_socio_8h or config.precio_socio_4h),
            # compatibilidad
            'price_particular': int(config.precio_particular_4h),
            'price_socio': int(config.precio_socio_4h),
            'capacity': config.capacidad,
            'type': 'social',
            'images': get_salon_images(config.salon.nombre)
        })
    
    # Obtener servicios adicionales activos para mostrarlos en el formulario (excluir montajes)
    from .models import ServicioAdicional
    servicios_adicionales = ServicioAdicional.objects.filter(activo=True).exclude(nombre__icontains='montaje').order_by('nombre')

    if request.method == 'GET':
        return render(request, 'register.html', {'rooms': rooms, 'servicios_adicionales': servicios_adicionales})
    
    # Procesar POST
    nombre_cliente = request.POST.get('nombre_cliente', '').strip()
    email_cliente = request.POST.get('email_cliente', '').strip()
    telefono_cliente = request.POST.get('telefono_cliente', '').strip()
    num_personas = request.POST.get('num_personas', '').strip()
    fecha_evento = request.POST.get('fecha_evento', '').strip()
    duracion_horas = request.POST.get('duracion_horas', '').strip()
    hora_inicio = request.POST.get('hora_inicio', '').strip()
    tiempo_decoracion = request.POST.get('tiempo_decoracion', '0').strip()
    espacio_id = request.POST.get('espacio_id', '0')
    socio_code = request.POST.get('socio_code', '').strip()
    note = request.POST.get('note', '').strip()
    
    # Validaciones
    errors = []
    if not nombre_cliente or len(nombre_cliente) < 2:
        errors.append("Nombre del cliente inválido.")
    if not email_cliente or '@' not in email_cliente:
        errors.append("Email inválido.")
    if not telefono_cliente or len(telefono_cliente) < 7:
        errors.append("Teléfono inválido.")
    
    try:
        personas_i = int(num_personas)
        if personas_i <= 0:
            errors.append("Número de personas inválido.")
    except:
        errors.append("Número de personas inválido.")
        personas_i = 0
    
    try:
        duracion_i = int(duracion_horas)
        if duracion_i <= 0:
            errors.append("Duración del evento inválida.")
    except:
        errors.append("Duración del evento inválida.")
        duracion_i = 0
    
    # Validar fecha del evento
    try:
        from datetime import date
        fecha_evento_obj = datetime.strptime(fecha_evento, '%Y-%m-%d').date()
        if fecha_evento_obj < date.today():
            errors.append("La fecha del evento no puede ser en el pasado.")
    except ValueError:
        errors.append("Formato de fecha inválido.")
        fecha_evento_obj = None
    
    try:
        from .models import ConfiguracionSalon, BloqueoEspacio
        espacio_id_int = int(espacio_id)
        configuracion = ConfiguracionSalon.objects.filter(id=espacio_id_int, salon__disponible=True).select_related('salon').first()
        if not configuracion:
            errors.append("Espacio seleccionado inválido.")
        else:
            # Validar que el salón no esté bloqueado en esa fecha
            if fecha_evento_obj:
                bloqueos = BloqueoEspacio.objects.filter(
                    salon=configuracion.salon,
                    activo=True,
                    fecha_inicio__lte=fecha_evento_obj,
                    fecha_fin__gte=fecha_evento_obj
                )
                if bloqueos.exists():
                    bloqueo = bloqueos.first()
                    fecha_inicio_str = bloqueo.fecha_inicio.strftime('%d/%m/%Y')
                    fecha_fin_str = bloqueo.fecha_fin.strftime('%d/%m/%Y')
                    errors.append(f"El salón {configuracion.salon.nombre} no está disponible para la fecha {fecha_evento}. Motivo: {bloqueo.get_motivo_display()}. Bloqueado desde {fecha_inicio_str} hasta {fecha_fin_str}.")
            
            # Validar capacidad del salón
            if personas_i > configuracion.capacidad:
                errors.append(f"El salón soporta máximo {configuracion.capacidad} personas. Solicitaste {personas_i}.")
            
            # Validar código de socio si se marcó como socio
            es_socio = request.POST.get('es_socio', 'no')
            if es_socio == 'si' and not socio_code:
                errors.append("Debes ingresar el código de socio.")
            elif es_socio == 'si' and not is_valid_socio(socio_code):
                errors.append("El código de socio ingresado no es válido (mínimo 3 caracteres).")
    except:
        errors.append("Espacio seleccionado inválido.")
        configuracion = None
    
    if errors:
        return render(request, 'register.html', {'rooms': rooms, 'errors': errors, 'form': request.POST}, status=400)
    
    # Determine price depending on socio code and duration (4H/8H)
    es_socio = request.POST.get('es_socio', 'no')
    if configuracion:
        applied_socio = False
        if es_socio == 'si' and is_valid_socio(socio_code):
            applied_socio = True

        # Precio según duración seleccionada
        if duracion_i == 8:
            if applied_socio:
                price = int(configuracion.precio_socio_8h or configuracion.precio_socio_4h)
            else:
                price = int(configuracion.precio_particular_8h or configuracion.precio_particular_4h)
        else:
            # por defecto 4 horas
            if applied_socio:
                price = int(configuracion.precio_socio_4h)
            else:
                price = int(configuracion.precio_particular_4h)

        espacio_nombre = f"{configuracion.salon.nombre} - {configuracion.get_tipo_configuracion_display()}"
    else:
        price = 0
        applied_socio = False
        espacio_nombre = ''

    total_price = price
    
    # Procesar tiempo de decoración
    try:
        tiempo_decoracion_i = int(tiempo_decoracion)
    except:
        tiempo_decoracion_i = 0
    
    # Guardar la reserva en la base de datos
    try:
        from datetime import time as datetime_time
        # Convertir hora_inicio string a objeto time
        hora_inicio_obj = None
        if hora_inicio:
            try:
                hora_parts = hora_inicio.split(':')
                hora_inicio_obj = datetime_time(int(hora_parts[0]), int(hora_parts[1]))
            except:
                pass
        
        # Obtener nombre de entidad si viene de empresa
        nombre_entidad = request.POST.get('nombre_entidad', '').strip()
        
        reserva = Reserva.objects.create(
            configuracion_salon=configuracion,
            nombre_cliente=nombre_cliente,
            email_cliente=email_cliente,
            telefono_cliente=telefono_cliente,
            tipo_cliente='SOCIO' if applied_socio else 'PARTICULAR',
            nombre_entidad=nombre_entidad if nombre_entidad else None,
            fecha_evento=fecha_evento,
            hora_inicio=hora_inicio_obj,
            duracion='4H' if duracion_i == 4 else '8H',
            tiempo_decoracion=tiempo_decoracion_i,
            numero_personas=personas_i,
            precio_total=0,
            estado='PENDIENTE',
            observaciones=note if note else ''
        )
        # Después de crear la reserva, dejar que el modelo calcule el precio si corresponde
        reserva.refresh_from_db()
        total_price = int(reserva.precio_total or 0)

        # Mensaje de éxito
        messages.success(request, f'¡Reserva #{reserva.id} creada exitosamente! Total: ${total_price} USD')

        # Preparar payload para mostrar al usuario
        abono_50 = int(total_price * 0.5)
        payload = {
            'reserva_id': reserva.id,
            'nombre_cliente': nombre_cliente,
            'email_cliente': email_cliente,
            'telefono_cliente': telefono_cliente,
            'num_personas': personas_i,
            'fecha_evento': fecha_evento,
            'hora_inicio': hora_inicio if hora_inicio else 'No especificada',
            'duracion_horas': duracion_i,
            'espacio_id': espacio_id_int,
            'espacio_nombre': espacio_nombre,
            'precio_por_evento': price,
            'total_price': total_price,
            'abono_50': f"{abono_50:,}",
            'socio_code': socio_code,
            'applied_socio': applied_socio,
            'note': note
        }
        
        return render(request, 'register.html', {'success': True, 'registration': payload, 'rooms': rooms})
    
    except Exception as e:
        messages.error(request, f'Error al crear la reserva: {str(e)}')
        errors.append(f"Error al guardar la reserva: {str(e)}")
        return render(request, 'register.html', {'errors': errors, 'rooms': rooms})

@staff_member_required
def admin(request):
    """Panel administrativo avanzado con filtros y métricas"""
    qs = Reserva.objects.select_related('configuracion_salon','configuracion_salon__salon').all()

    # Cambio de estado rápido via POST
    if request.method == 'POST' and request.user.is_staff:
        reserva_id = request.POST.get('reserva_id')
        nuevo_estado = request.POST.get('nuevo_estado')
        if reserva_id and nuevo_estado in ['PENDIENTE','CONFIRMADA','CANCELADA','COMPLETADA']:
            try:
                r = qs.get(id=reserva_id)
                estado_anterior = r.estado
                r.estado = nuevo_estado
                r.save()
                messages.success(request, f'Reserva #{r.id} cambió de {estado_anterior} a {nuevo_estado}')
            except Reserva.DoesNotExist:
                messages.error(request, 'Reserva no encontrada')

    # Filtros GET
    estado = request.GET.get('estado')
    tipo = request.GET.get('tipo')  # SOCIO / PARTICULAR
    salon = request.GET.get('salon')  # nombre salon
    desde = request.GET.get('desde')  # yyyy-mm-dd
    hasta = request.GET.get('hasta')

    if estado:
        qs = qs.filter(estado=estado)
    if tipo:
        qs = qs.filter(tipo_cliente=tipo)
    if salon:
        qs = qs.filter(configuracion_salon__salon__nombre__icontains=salon)
    if desde:
        qs = qs.filter(fecha_evento__gte=desde)
    if hasta:
        qs = qs.filter(fecha_evento__lte=hasta)

    qs = qs.order_by('-fecha_creacion')

    # Métricas
    total_reservas = qs.count()
    pendientes = qs.filter(estado='PENDIENTE').count()
    confirmadas = qs.filter(estado='CONFIRMADA').count()
    canceladas = qs.filter(estado='CANCELADA').count()
    completadas = qs.filter(estado='COMPLETADA').count()

    # Ingresos estimados por estado (solo confirmadas + completadas)
    ingresos = qs.filter(estado__in=['CONFIRMADA','COMPLETADA']).aggregate(total=Sum('precio_total'))['total'] or 0

    # Distribución por salón
    salas_stats = (
        qs.values('configuracion_salon__salon__nombre')
          .annotate(cantidad=Count('id'), monto=Sum('precio_total'))
          .order_by('-cantidad')[:10]
    )

    context = {
        'reservas': qs,
        'total_reservas': total_reservas,
        'pendientes': pendientes,
        'confirmadas': confirmadas,
        'canceladas': canceladas,
        'completadas': completadas,
        'ingresos': ingresos,
        'salas_stats': salas_stats,
        'estados': ['PENDIENTE','CONFIRMADA','CANCELADA','COMPLETADA'],
        'f_estado': estado or '',
        'f_tipo': tipo or '',
        'f_salon': salon or '',
        'f_desde': desde or '',
        'f_hasta': hasta or '',
    }

    return render(request, 'admin.html', context)

@login_required
def logout_view(request):
    """Logout seguro solo vía POST para evitar 405 en algunos navegadores y extensiones."""
    if request.method == 'POST':
        logout(request)
        messages.info(request, 'Sesión cerrada correctamente')
    return redirect('reservas:index')

@staff_member_required
def borrar_reservas(request):
    """Borrar todas las reservas (solo staff)"""
    if request.method == 'POST':
        count = Reserva.objects.count()
        Reserva.objects.all().delete()
        messages.warning(request, f'Se eliminaron {count} reserva(s) correctamente')
        return redirect('reservas:panel')
    return redirect('reservas:panel')

@staff_member_required
def export_reservas_csv(request):
    """Exportar reservas a Excel con análisis de ganancias"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.db.models import Sum, Count, Q
        from django.db.models.functions import TruncMonth
    except ImportError:
        # Fallback a CSV si openpyxl no está instalado
        return export_reservas_csv_fallback(request)
    
    qs = Reserva.objects.select_related('configuracion_salon','configuracion_salon__salon').all()
    
    # Aplicar filtros GET
    estado = request.GET.get('estado')
    tipo = request.GET.get('tipo')
    salon = request.GET.get('salon')
    desde = request.GET.get('desde')
    hasta = request.GET.get('hasta')
    
    if estado:
        qs = qs.filter(estado=estado)
    if tipo:
        qs = qs.filter(tipo_cliente=tipo)
    if salon:
        qs = qs.filter(configuracion_salon__salon__nombre__icontains=salon)
    if desde:
        qs = qs.filter(fecha_evento__gte=desde)
    if hasta:
        qs = qs.filter(fecha_evento__lte=hasta)
    
    # Crear workbook
    wb = openpyxl.Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    total_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # HOJA 1: Listado de Reservas
    ws1 = wb.active
    ws1.title = "Reservas"
    
    headers = ['ID', 'Cliente', 'Email', 'Teléfono', 'Entidad/Empresa', 'Salón', 'Configuración', 
               'Fecha Evento', 'Hora Inicio', 'Duración', 'Personas', 'Tiempo Decoración', 
               'Tipo Cliente', 'Precio Total', 'Estado', 'Fecha Reserva', 'Observaciones']
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(1, col, header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row = 2
    total_ingresos = 0
    for r in qs:
        ws1.cell(row, 1, r.id)
        ws1.cell(row, 2, r.nombre_cliente)
        ws1.cell(row, 3, r.email_cliente)
        ws1.cell(row, 4, r.telefono_cliente)
        ws1.cell(row, 5, r.nombre_entidad or 'N/A')
        ws1.cell(row, 6, r.configuracion_salon.salon.nombre)
        ws1.cell(row, 7, r.configuracion_salon.get_tipo_configuracion_display())
        ws1.cell(row, 8, r.fecha_evento.strftime('%Y-%m-%d'))
        ws1.cell(row, 9, r.hora_inicio.strftime('%H:%M') if r.hora_inicio else 'N/A')
        ws1.cell(row, 10, r.get_duracion_display())
        ws1.cell(row, 11, r.numero_personas)
        ws1.cell(row, 12, f"{r.tiempo_decoracion}h")
        ws1.cell(row, 13, r.get_tipo_cliente_display())
        ws1.cell(row, 14, float(r.precio_total))
        ws1.cell(row, 15, r.get_estado_display())
        ws1.cell(row, 16, r.fecha_creacion.strftime('%Y-%m-%d %H:%M'))
        ws1.cell(row, 17, r.observaciones or '')
        
        if r.estado in ['CONFIRMADA', 'COMPLETADA']:
            total_ingresos += float(r.precio_total)
        
        for col in range(1, 18):
            ws1.cell(row, col).border = border
        
        row += 1
    
    # Total al final
    ws1.cell(row, 13, "TOTAL INGRESOS:").font = total_font
    ws1.cell(row, 14, total_ingresos).font = total_font
    ws1.cell(row, 14).number_format = '$#,##0.00'
    ws1.cell(row, 13).fill = total_fill
    ws1.cell(row, 14).fill = total_fill
    
    # Ajustar anchos
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 30
    ws1.column_dimensions['D'].width = 15
    ws1.column_dimensions['E'].width = 25
    ws1.column_dimensions['F'].width = 20
    ws1.column_dimensions['G'].width = 15
    ws1.column_dimensions['H'].width = 15
    ws1.column_dimensions['I'].width = 12
    ws1.column_dimensions['J'].width = 12
    ws1.column_dimensions['K'].width = 10
    ws1.column_dimensions['L'].width = 12
    ws1.column_dimensions['M'].width = 15
    ws1.column_dimensions['N'].width = 15
    ws1.column_dimensions['O'].width = 12
    ws1.column_dimensions['P'].width = 18
    ws1.column_dimensions['Q'].width = 30
    
    # HOJA 2: Resumen por Mes
    ws2 = wb.create_sheet("Resumen Mensual")
    
    # Headers
    ws2.cell(1, 1, "Mes").fill = header_fill
    ws2.cell(1, 1).font = header_font
    ws2.cell(1, 2, "Total Reservas").fill = header_fill
    ws2.cell(1, 2).font = header_font
    ws2.cell(1, 3, "Confirmadas").fill = header_fill
    ws2.cell(1, 3).font = header_font
    ws2.cell(1, 4, "Ingresos").fill = header_fill
    ws2.cell(1, 4).font = header_font
    
    # Agrupar por mes
    reservas_por_mes = Reserva.objects.filter(
        Q(estado='CONFIRMADA') | Q(estado='COMPLETADA')
    ).annotate(
        mes=TruncMonth('fecha_evento')
    ).values('mes').annotate(
        total=Count('id'),
        ingresos=Sum('precio_total')
    ).order_by('mes')
    
    row = 2
    total_mes = 0
    for item in reservas_por_mes:
        ws2.cell(row, 1, item['mes'].strftime('%B %Y'))
        ws2.cell(row, 2, item['total'])
        ws2.cell(row, 3, item['total'])
        ws2.cell(row, 4, float(item['ingresos'] or 0))
        ws2.cell(row, 4).number_format = '$#,##0.00'
        total_mes += float(item['ingresos'] or 0)
        row += 1
    
    # Total anual
    ws2.cell(row, 3, "TOTAL AÑO:").font = total_font
    ws2.cell(row, 4, total_mes).font = total_font
    ws2.cell(row, 4).number_format = '$#,##0.00'
    ws2.cell(row, 3).fill = total_fill
    ws2.cell(row, 4).fill = total_fill
    
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 20
    
    # HOJA 3: Resumen por Salón
    ws3 = wb.create_sheet("Resumen por Salón")
    
    ws3.cell(1, 1, "Salón").fill = header_fill
    ws3.cell(1, 1).font = header_font
    ws3.cell(1, 2, "Total Reservas").fill = header_fill
    ws3.cell(1, 2).font = header_font
    ws3.cell(1, 3, "Ingresos").fill = header_fill
    ws3.cell(1, 3).font = header_font
    
    reservas_por_salon = Reserva.objects.filter(
        Q(estado='CONFIRMADA') | Q(estado='COMPLETADA')
    ).values(
        'configuracion_salon__salon__nombre'
    ).annotate(
        total=Count('id'),
        ingresos=Sum('precio_total')
    ).order_by('-ingresos')
    
    row = 2
    for item in reservas_por_salon:
        ws3.cell(row, 1, item['configuracion_salon__salon__nombre'])
        ws3.cell(row, 2, item['total'])
        ws3.cell(row, 3, float(item['ingresos'] or 0))
        ws3.cell(row, 3).number_format = '$#,##0.00'
        row += 1
    
    ws3.column_dimensions['A'].width = 25
    ws3.column_dimensions['B'].width = 15
    ws3.column_dimensions['C'].width = 20
    
    # Guardar en respuesta
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reservas_club_el_meta_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    return response


def export_reservas_csv_fallback(request):
    """Fallback CSV si openpyxl no está disponible"""
    qs = Reserva.objects.select_related('configuracion_salon','configuracion_salon__salon').all()
    
    estado = request.GET.get('estado')
    tipo = request.GET.get('tipo')
    salon = request.GET.get('salon')
    desde = request.GET.get('desde')
    hasta = request.GET.get('hasta')
    
    if estado:
        qs = qs.filter(estado=estado)
    if tipo:
        qs = qs.filter(tipo_cliente=tipo)
    if salon:
        qs = qs.filter(configuracion_salon__salon__nombre__icontains=salon)
    if desde:
        qs = qs.filter(fecha_evento__gte=desde)
    if hasta:
        qs = qs.filter(fecha_evento__lte=hasta)
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="reservas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow(['ID', 'Cliente', 'Email', 'Teléfono', 'Salón', 'Configuración', 'Fecha Evento', 
                     'Personas', 'Duración', 'Tipo Cliente', 'Precio Total', 'Estado', 'Fecha Reserva', 'Observaciones'])
    
    for r in qs:
        writer.writerow([
            r.id,
            r.nombre_cliente,
            r.email_cliente,
            r.telefono_cliente,
            r.configuracion_salon.salon.nombre,
            r.configuracion_salon.get_tipo_configuracion_display(),
            r.fecha_evento,
            r.numero_personas,
            r.get_duracion_display(),
            r.get_tipo_cliente_display(),
            r.precio_total,
            r.get_estado_display(),
            r.fecha_creacion.strftime('%Y-%m-%d %H:%M:%S'),
            r.observaciones or ''
        ])
    
    return response


def check_availability(request):
    """API endpoint para verificar disponibilidad de espacios en una fecha específica"""
    fecha_str = request.GET.get('fecha', '')
    
    try:
        fecha_obj = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except:
        return JsonResponse({'error': 'Fecha inválida'}, status=400)
    
    # Obtener todas las configuraciones
    configuraciones = ConfiguracionSalon.objects.filter(salon__disponible=True).select_related('salon')
    
    # Verificar cuáles salones están bloqueados
    salones_bloqueados = set()
    bloqueos = BloqueoEspacio.objects.filter(
        activo=True,
        fecha_inicio__lte=fecha_obj,
        fecha_fin__gte=fecha_obj
    ).select_related('salon')
    
    bloqueos_info = {}
    for bloqueo in bloqueos:
        salones_bloqueados.add(bloqueo.salon.id)
        bloqueos_info[bloqueo.salon.id] = {
            'motivo': bloqueo.get_motivo_display(),
            'descripcion': bloqueo.descripcion
        }
    
    # Construir respuesta
    result = []
    for config in configuraciones:
        esta_bloqueado = config.salon.id in salones_bloqueados
        item = {
            'id': config.id,
            'salon_id': config.salon.id,
            'salon_nombre': config.salon.nombre,
            'configuracion': config.get_tipo_configuracion_display(),
            'capacidad': config.capacidad,
            'disponible': not esta_bloqueado,
        }
        
        if esta_bloqueado:
            item['motivo_bloqueo'] = bloqueos_info[config.salon.id]['motivo']
            item['descripcion_bloqueo'] = bloqueos_info[config.salon.id]['descripcion']
        
        result.append(item)
    
    return JsonResponse({'espacios': result, 'fecha': fecha_str})


def get_bloqueos_salon(request):
    """API endpoint para obtener todos los bloqueos de un salón específico"""
    salon_id = request.GET.get('salon_id', '')
    
    if not salon_id:
        return JsonResponse({'error': 'salon_id requerido'}, status=400)
    
    try:
        # Obtener bloqueos activos del salón
        bloqueos = BloqueoEspacio.objects.filter(
            salon_id=salon_id,
            activo=True
        ).order_by('fecha_inicio')
        
        result = []
        for bloqueo in bloqueos:
            result.append({
                'fecha_inicio': bloqueo.fecha_inicio.strftime('%Y-%m-%d'),
                'fecha_fin': bloqueo.fecha_fin.strftime('%Y-%m-%d'),
                'motivo': bloqueo.get_motivo_display(),
                'descripcion': bloqueo.descripcion
            })
        
        return JsonResponse({'bloqueos': result})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def reportes_dashboard(request):
    """Vista principal del dashboard de reportes avanzados"""
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para acceder a los reportes.')
        return redirect('index')
    
    # Obtener parámetros de filtro
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    espacio_id = request.GET.get('espacio_id')
    preset = request.GET.get('preset')  # '30d', 'month', 'year', 'custom'

    # Valores por defecto dependiendo del preset
    hoy = datetime.now().date()
    if preset and preset != 'custom':
        import calendar
        def subtract_months(d, months):
            m = d.month - months
            y = d.year
            while m <= 0:
                m += 12
                y -= 1
            last_day = calendar.monthrange(y, m)[1]
            day = min(d.day, last_day)
            return d.replace(year=y, month=m, day=day)

        if preset == '7d':
            fecha_fin_obj = hoy
            fecha_inicio_obj = fecha_fin_obj - timedelta(days=7)
        elif preset == '15d':
            fecha_fin_obj = hoy
            fecha_inicio_obj = fecha_fin_obj - timedelta(days=15)
        elif preset == '30d':
            fecha_fin_obj = hoy
            fecha_inicio_obj = fecha_fin_obj - timedelta(days=30)
        elif preset == '90d':
            fecha_fin_obj = hoy
            fecha_inicio_obj = fecha_fin_obj - timedelta(days=90)
        elif preset == '6m':
            fecha_fin_obj = hoy
            fecha_inicio_obj = subtract_months(fecha_fin_obj, 6)
        elif preset == 'month':
            # Mes actual: incluir hasta el último día del mes
            first_day = date(hoy.year, hoy.month, 1)
            last_day_num = (first_day.replace(day=28) + timedelta(days=4)).day
            # compute last day correctly
            import calendar
            last_day_num = calendar.monthrange(hoy.year, hoy.month)[1]
            fecha_inicio_obj = first_day
            fecha_fin_obj = date(hoy.year, hoy.month, last_day_num)
        elif preset == 'year':
            fecha_inicio_obj = date(hoy.year, 1, 1)
            fecha_fin_obj = date(hoy.year, 12, 31)
        else:
            fecha_fin_obj = hoy
            fecha_inicio_obj = fecha_fin_obj - timedelta(days=30)
        fecha_inicio = fecha_inicio_obj.strftime('%Y-%m-%d')
        fecha_fin = fecha_fin_obj.strftime('%Y-%m-%d')
    else:
        # Sin preset o preset personalizado: usar parámetros si vienen o fallback últimos 30 días
        if not fecha_fin:
            fecha_fin_obj = hoy
            fecha_fin = fecha_fin_obj.strftime('%Y-%m-%d')
        else:
            try:
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            except:
                fecha_fin_obj = hoy
                fecha_fin = fecha_fin_obj.strftime('%Y-%m-%d')

        if not fecha_inicio:
            fecha_inicio_obj = fecha_fin_obj - timedelta(days=30)
            fecha_inicio = fecha_inicio_obj.strftime('%Y-%m-%d')
        else:
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            except:
                fecha_inicio_obj = fecha_fin_obj - timedelta(days=30)
                fecha_inicio = fecha_inicio_obj.strftime('%Y-%m-%d')
    
    # Query base de reservas
    reservas_query = Reserva.objects.filter(
        estado='confirmada',
        fecha__gte=fecha_inicio_obj,
        fecha__lte=fecha_fin_obj
    ).select_related('configuracion_salon__salon')
    
    # Filtrar por espacio si se especifica
    if espacio_id:
        reservas_query = reservas_query.filter(configuracion_salon__salon_id=espacio_id)
    
    reservas = reservas_query.all()
    
    # MÉTRICAS PRINCIPALES
    total_reservas = reservas.count()
    
    # Calcular ingresos totales
    ingresos_totales = 0
    for reserva in reservas:
        ingresos_totales += reserva.precio_total or 0
    
    # Reservas por tipo de cliente
    reservas_socio = reservas.filter(tipo_cliente='socio').count()
    reservas_particular = reservas.filter(tipo_cliente='particular').count()
    
    # Promedio de personas por reserva
    total_personas = sum([r.numero_personas or 0 for r in reservas])
    promedio_personas = round(total_personas / total_reservas, 1) if total_reservas > 0 else 0
    
    # ANÁLISIS POR ESPACIO
    espacios_stats = {}
    for reserva in reservas:
        salon = reserva.configuracion_salon.salon
        if salon.id not in espacios_stats:
            espacios_stats[salon.id] = {
                'nombre': salon.nombre,
                'reservas': 0,
                'ingresos': 0,
                'personas': 0
            }
        espacios_stats[salon.id]['reservas'] += 1
        espacios_stats[salon.id]['ingresos'] += reserva.precio_total or 0
        espacios_stats[salon.id]['personas'] += reserva.numero_personas or 0
    
    # Convertir a lista ordenada por ingresos
    espacios_ranking = sorted(
        espacios_stats.values(),
        key=lambda x: x['ingresos'],
        reverse=True
    )
    
    # INGRESOS POR DÍA (para gráfica)
    ingresos_por_dia = {}
    reservas_por_dia = {}
    
    current_date = fecha_inicio_obj
    while current_date <= fecha_fin_obj:
        date_str = current_date.strftime('%Y-%m-%d')
        ingresos_por_dia[date_str] = 0
        reservas_por_dia[date_str] = 0
        current_date += timedelta(days=1)
    
    for reserva in reservas:
        date_str = reserva.fecha.strftime('%Y-%m-%d')
        if date_str in ingresos_por_dia:
            ingresos_por_dia[date_str] += reserva.precio_total or 0
            reservas_por_dia[date_str] += 1
    
    # Preparar datos para gráficas
    fechas_labels = list(ingresos_por_dia.keys())
    ingresos_values = list(ingresos_por_dia.values())
    reservas_values = list(reservas_por_dia.values())
    
    # SERVICIOS ADICIONALES
    servicios_totales = ReservaServicioAdicional.objects.filter(
        reserva__in=reservas
    ).select_related('servicio')
    
    ingresos_servicios = sum([s.subtotal for s in servicios_totales])
    
    servicios_stats = {}
    for servicio_reserva in servicios_totales:
        servicio = servicio_reserva.servicio.nombre
        if servicio not in servicios_stats:
            servicios_stats[servicio] = {
                'cantidad': 0,
                'ingresos': 0
            }
        servicios_stats[servicio]['cantidad'] += servicio_reserva.cantidad
        servicios_stats[servicio]['ingresos'] += servicio_reserva.subtotal
    
    # Obtener todos los espacios para el filtro
    espacios = Salon.objects.all().order_by('nombre')
    
    context = {
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'espacio_id': espacio_id,
        'espacios': espacios,
        'total_reservas': total_reservas,
        'ingresos_totales': ingresos_totales,
        'ingresos_servicios': ingresos_servicios,
        'ingresos_gran_total': ingresos_totales + ingresos_servicios,
        'reservas_socio': reservas_socio,
        'reservas_particular': reservas_particular,
        'promedio_personas': promedio_personas,
        'espacios_ranking': espacios_ranking,
        'servicios_stats': servicios_stats,
        'fechas_labels': fechas_labels,
        'ingresos_values': ingresos_values,
        'reservas_values': reservas_values,
        'preset': preset or '30d',
    }
    
    return render(request, 'reportes.html', context)
